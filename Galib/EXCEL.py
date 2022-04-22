"""Импорты"""
import os
from traceback import print_tb, print_exc
import sys
# from Galib.EXСEPTION_HANDLER import ExcelopenError, ExcelwriteError, ExcelsaveError
import openpyxl
from config import Config as cfg
from Galib.EXСEPTION_HANDLER import Excelcheck_recordError, ExcelsaveError, ExcelwriteError, DataframeError
from traceback import print_tb, print_exc


class Excel:

    @staticmethod
    def check_record(file):
        """Проверяет наличие записей в файле"""
        try:
            workbook = os.path.join(cfg.folder_input, file)
            wb = openpyxl.load_workbook(workbook)
            sheet = wb['Sheet1']
            if sheet.cell(row=2, column=1).value:
                return True
            else:
                return False
        except Exception:
            print_exc(limit=2, file=sys.stdout)
            raise Excelcheck_recordError('Ошибка в блоке Excel.check_record')


    def writer(self, file, name_column, value):
        """Сохраняет значение в нужный столбец"""
        try:
            wb = openpyxl.load_workbook(file)
            sheet = wb['Sheet1']
            column = self.search_column(wb, name_column)
            row = self.end_row(wb, column)
            sheet.cell(row=row, column=column).value = value
            wb.save(file)
            wb.close()
        except Exception:
            print_exc(limit=2, file=sys.stdout)
            raise ExcelwriteError('Ошибка в блоке Excel.writer')


    def end_row(self, file, column):
        """Возвращает индекс первой пустой строки"""
        wb = file
        sheet = wb['Sheet1']
        row = 2
        while True:
            if sheet.cell(row=row, column=column).value:
                row += 1
                continue
            return row

    def search_column(self, file, column):
        """Возвращает индекс столбца"""
        wb = file
        sheet = wb['Sheet1']
        for index in range(1, sheet.max_column + 1):
            value_column = sheet.cell(row=1, column=index).value
            if column in value_column:
                return index




if __name__ == "__main__":
    file = '/home/serj52/PycharmProjects/search_on_rambler/OUTPUT/output.xlsx'
    e = Excel()
    e.writer(file, 'Ссылка', 'https')
    e.writer(file, 'Наименование материала', 'гайка')
    e.writer(file, 'Ссылка', 'https')
    e.writer(file, 'Наименование материала', 'болт')